import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, IconSetRule


# ==========================
# CONVERSÃO DE VALORES
# ==========================
def converter_valor(valor_str):
    if valor_str is None:
        return 0.0

    valor_str = str(valor_str).strip()
    if valor_str == "":
        return 0.0

    valor_str = ''.join(c for c in valor_str if c.isdigit() or c in [',', '.', '-'])

    if valor_str in ("", "-"):
        return 0.0

    if ',' in valor_str:
        valor_str = valor_str.replace('.', '').replace(',', '.')
    else:
        if valor_str.count('.') > 1:
            valor_str = valor_str.replace('.', '')

    try:
        return float(valor_str)
    except ValueError:
        return 0.0


# ==========================
# LEITURA DO CSV
# ==========================
def ler_csv(nome_arquivo):
    encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']
    dados = []

    for encoding in encodings:
        try:
            with open(nome_arquivo, encoding=encoding) as f:
                reader = csv.DictReader(f, delimiter=';')

                for row in reader:
                    dados.append({
                        'EQUIPE': row.get('EQUIPE', '').strip(),
                        'VEND': row.get('VEND', '').strip(),
                        'AREA': row.get('AREA', '').strip(),
                        'FORNECEDOR': row.get('FORNECEDOR', '').strip(),
                        'TIPO': row.get('TIPO', '').strip().upper(),
                        'VALOR': converter_valor(row.get('VALOR'))
                    })

                print(f"CSV lido com encoding: {encoding}")
                return dados

        except UnicodeDecodeError:
            continue

    raise Exception("Não foi possível ler o CSV.")


# ==========================
# GERAÇÃO DO EXCEL
# ==========================
def gerar_excel(resultados, fornecedores, nome_arquivo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Metas Distribuídas"

    fornecedores = sorted(fornecedores)

    # Normaliza o retorno do motor ({fornecedor: [entidades]})
    # para a estrutura por linha ({(equipe, vendedor): {fornecedor: dados}})
    linhas = {}
    for fornecedor in fornecedores:
        entidades = resultados.get(fornecedor, []) or []
        for entidade in entidades:
            equipe = entidade.get("equipe", "")
            vendedor = entidade.get("vendedor", "")
            chave = (equipe, vendedor)

            if chave not in linhas:
                linhas[chave] = {}

            linhas[chave][fornecedor] = entidade

    chaves = sorted(linhas.keys())

    ws["A1"] = "EQUIPE"
    ws["B1"] = "VENDEDOR"

    col = 3
    mapa_colunas = {}

    # ---------- CABEÇALHO ----------
    for fornecedor in fornecedores:
        mapa_colunas[fornecedor] = col

        ws.cell(row=1, column=col,     value=f"{fornecedor} - Média")
        ws.cell(row=1, column=col + 1, value=f"{fornecedor} - Histórico")
        ws.cell(row=1, column=col + 2, value=f"{fornecedor} - Meta")
        ws.cell(row=1, column=col + 3, value=f"{fornecedor} - Índice Pressão")

        col += 4

    # ---------- ESTILO CABEÇALHO ----------
    for c in range(1, col):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="366092")
        cell.alignment = Alignment(horizontal="center")

    # ---------- DADOS ----------
    linha = 2
    for equipe, vendedor in chaves:
        ws.cell(row=linha, column=1, value=equipe)
        ws.cell(row=linha, column=2, value=vendedor)

        for fornecedor in fornecedores:
            base_col = mapa_colunas[fornecedor]
            dados = linhas[(equipe, vendedor)].get(fornecedor, {})
            meta = dados.get("meta_final", dados.get("meta_distribuida", 0))

            ws.cell(row=linha, column=base_col,     value=dados.get("media", 0))
            ws.cell(row=linha, column=base_col + 1, value=dados.get("historico", 0))
            ws.cell(row=linha, column=base_col + 2, value=meta)
            ws.cell(row=linha, column=base_col + 3, value=dados.get("indice_pressao", 0))

        linha += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    for c in range(3, col):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # ==========================
    # FORMATAÇÃO CONDICIONAL (SETAS)
    # ==========================
    ultima_linha = linha - 1

    for fornecedor in fornecedores:
        col_indice = mapa_colunas[fornecedor] + 3
        letra = get_column_letter(col_indice)
        intervalo = f"{letra}2:{letra}{ultima_linha}"

        # 1️⃣ Não formatar se índice = 0 (sem meta)
        regra_bloqueio = FormulaRule(
            formula=[f"{letra}2=0"],
            stopIfTrue=True
        )

        ws.conditional_formatting.add(intervalo, regra_bloqueio)

        # 2️⃣ Setas (quanto maior o índice, pior)
        regra_setas = IconSetRule(
            icon_style="3Arrows",
            type="num",
            values=[1, 1.2],
            showValue=True,
            reverse=False
        )

        ws.conditional_formatting.add(intervalo, regra_setas)


    wb.save(nome_arquivo)
    print(f"Arquivo Excel gerado: {nome_arquivo}")
